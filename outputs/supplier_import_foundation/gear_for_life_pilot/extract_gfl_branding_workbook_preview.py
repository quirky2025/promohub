import csv
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


SOURCE_XLSX = Path(r"C:\Users\jilin\Desktop\supplier\gearforlife\The Source Branding Price List - Combined.xlsx")
OUT_DIR = Path(r"C:\Users\jilin\Desktop\promohub\outputs\supplier_import_foundation\gear_for_life_pilot")

OPTIONS_CSV = OUT_DIR / "gear_for_life_branding_options_FROM_XLSX_PREVIEW.csv"
PRICE_ROWS_CSV = OUT_DIR / "gear_for_life_branding_price_rows_FROM_XLSX_PREVIEW.csv"
GENERAL_RATE_ROWS_CSV = OUT_DIR / "gear_for_life_branding_general_rate_card_rows_FROM_XLSX_PREVIEW.csv"
GENERAL_RATE_CARDS_CSV = OUT_DIR / "gear_for_life_branding_general_rate_cards_FROM_XLSX_PREVIEW.csv"
AUDIT_MD = OUT_DIR / "gear_for_life_branding_workbook_audit.md"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_money(value):
    text = clean(value)
    if not text:
        return ""
    if text.upper() == "POA":
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    return f"{float(match.group(0)):.2f}" if match else ""


def price_status(value):
    text = clean(value)
    if not text:
        return ""
    upper = text.upper()
    if upper == "POA":
        return "poa"
    if "NOT APPLICABLE" in upper or "NOT AVAILABLE" in upper:
        return "unavailable"
    return "priced" if parse_money(text) else "request_quote"


def slug_part(value):
    text = clean(value).lower()
    text = text.replace("–", "-").replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def parse_setup(text):
    text = clean(text)
    if not text:
        return "", ""
    nums = re.findall(r"\$?\s*(\d+(?:\.\d+)?)", text.replace(",", ""))
    setup = f"{float(nums[0]):.2f}" if nums else ""
    repeat = f"{float(nums[1]):.2f}" if len(nums) > 1 else ""
    return setup, repeat


def parse_size(location):
    text = clean(location)
    match = re.search(r"(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*mm", text, re.I)
    if not match:
        return "", "", ""
    return f"{float(match.group(1)):.2f}", f"{float(match.group(2)):.2f}", f"{match.group(1)}x{match.group(2)}mm"


def parse_location_and_size(location):
    text = clean(location)
    if " - " in text:
        loc = text.split(" - ", 1)[0].strip()
    else:
        loc = re.split(r"\s+-\s+|\s+\(", text, 1)[0].strip()
    width, height, size = parse_size(text)
    return loc, size, width, height


def pricing_basis(method):
    m = clean(method).lower()
    if "pad" in m:
        return "per_colour_position"
    if "laser" in m:
        return "per_position"
    if "uvdtf" in m:
        return "per_unit"
    if "transfer" in m:
        return "per_unit"
    if "embroidery" in m:
        return "per_stitch_band"
    return "other"


def parse_qty_range(text):
    text = clean(text).replace(",", "")
    if not text:
        return "", ""
    if text.endswith("+"):
        return text[:-1], ""
    match = re.match(r"(\d+)\s*-\s*(\d+)", text)
    if match:
        return match.group(1), match.group(2)
    if text.isdigit():
        return text, text
    return "", ""


def parse_stitch_range(text):
    return parse_qty_range(text)


def option_key(sku, method, location):
    _, size, _, _ = parse_location_and_size(location)
    return "|".join([slug_part(sku), slug_part(method), slug_part(location), slug_part(size)])


def load_workbook_rows():
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["All Branding Prices"]
    return ws


def extract_product_specific_rows(ws):
    headers = [clean(cell.value) for cell in ws[4]]
    qty_cols = [(idx, int(headers[idx].replace("+", ""))) for idx in range(4, 10) if headers[idx].endswith("+")]

    current_sku = ""
    current_name = ""
    current_method = ""
    current_setup = ""
    options = {}
    price_rows = []
    unavailable_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=277, values_only=True), start=5):
        sku = clean(row[0])
        name = clean(row[1])
        method = clean(row[2])
        location = clean(row[3])
        setup_text = clean(row[10]) if len(row) > 10 else ""

        if sku:
            current_sku = sku
            current_name = name
            current_method = ""
            current_setup = ""
        if method:
            current_method = method
        if setup_text:
            current_setup = setup_text

        if not location:
            continue

        values = [clean(row[col_idx]) for col_idx, _ in qty_cols]
        if not any(values) and "POA" not in location.upper():
            continue

        loc, size_label, width, height = parse_location_and_size(location)
        setup_cost, repeat_setup_cost = parse_setup(current_setup)
        statuses = [price_status(v) for v in values if v]
        option_status = "priced" if "priced" in statuses else "poa" if "POA" in location.upper() or "poa" in statuses else "unavailable" if "unavailable" in statuses else "request_quote"
        key = option_key(current_sku, current_method, location)
        add_colour = "Additional colour print requirement POA" if "pad" in current_method.lower() else ""

        options[key] = {
            "supplier": "Gear For Life",
            "supplier_sku": current_sku,
            "product_name": current_name,
            "decoration_option_key": key,
            "decoration_method": current_method,
            "decoration_area": location,
            "decoration_location": loc,
            "artwork_size_label": size_label,
            "max_width_mm": width,
            "max_height_mm": height,
            "pricing_model": "sku_location_qty",
            "price_status": option_status,
            "setup_cost": setup_cost,
            "repeat_setup_cost": repeat_setup_cost,
            "setup_cost_label": current_setup,
            "pricing_basis": pricing_basis(current_method),
            "additional_colour_policy": add_colour,
            "source_row_number": row_idx,
            "raw_json": json.dumps(
                {
                    "source_row_number": row_idx,
                    "raw_location": location,
                    "raw_method": current_method,
                    "raw_setup": current_setup,
                },
                ensure_ascii=False,
            ),
        }

        for col_idx, qty in qty_cols:
            raw_price = clean(row[col_idx])
            status = price_status(raw_price)
            if not raw_price:
                continue
            unit_cost = parse_money(raw_price)
            price_rows.append(
                {
                    "supplier": "Gear For Life",
                    "supplier_sku": current_sku,
                    "product_name": current_name,
                    "decoration_option_key": key,
                    "decoration_method": current_method,
                    "decoration_area": location,
                    "decoration_location": loc,
                    "artwork_size_label": size_label,
                    "currency": "AUD",
                    "min_qty": qty,
                    "max_qty": "",
                    "unit_cost": unit_cost,
                    "setup_cost": setup_cost,
                    "repeat_setup_cost": repeat_setup_cost,
                    "pricing_basis": pricing_basis(current_method),
                    "price_status": status,
                    "raw_price": raw_price,
                    "source_row_number": row_idx,
                }
            )
            if status == "unavailable":
                unavailable_rows.append((row_idx, current_sku, location, raw_price))

    return list(options.values()), price_rows, unavailable_rows


def extract_general_rate_cards(ws):
    cards = [
        {
            "supplier": "Gear For Life",
            "rate_card_key": "transfer_printing_bags",
            "decoration_method": "Transfer Printing",
            "applies_to": "Bags",
            "applies_to_category": "Bags",
            "applies_to_subcategory": "",
            "is_default_for_scope": "true",
            "fallback_policy": "when_no_product_specific_option",
            "pricing_model": "size_qty_matrix",
            "setup_cost": "55.00",
            "repeat_setup_cost": "",
            "surcharge_cost": "",
            "surcharge_label": "",
            "notes": "Transfer Printing - Bags (Full Colour Transfer, priced by size). Default for bag products when no product-specific decoration method is supplied.",
        },
        {
            "supplier": "Gear For Life",
            "rate_card_key": "embroidery_apparel_selected_bags",
            "decoration_method": "Embroidery",
            "applies_to": "Apparel and selected bags",
            "applies_to_category": "",
            "applies_to_subcategory": "",
            "is_default_for_scope": "false",
            "fallback_policy": "manual_review",
            "pricing_model": "stitch_count_qty",
            "setup_cost": "75.00",
            "repeat_setup_cost": "",
            "surcharge_cost": "1.90",
            "surcharge_label": "Some items +$1.90/unit",
            "notes": "Embroidery by stitch count and quantity. Logo setup $75/logo.",
        },
    ]
    rows = []

    # Transfer printing matrix.
    transfer_setup = "55.00"
    transfer_sizes = [clean(v) for v in next(ws.iter_rows(min_row=282, max_row=282, values_only=True))[1:5]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=283, max_row=286, values_only=True), start=283):
        qty_min, qty_max = parse_qty_range(row[0])
        for offset, size in enumerate(transfer_sizes, start=1):
            raw_price = clean(row[offset])
            rows.append(
                {
                    "supplier": "Gear For Life",
                    "rate_card_key": "transfer_printing_bags",
                    "decoration_method": "Transfer Printing",
                    "applies_to": "Bags",
                    "artwork_size_label": size,
                    "stitch_count_min": "",
                    "stitch_count_max": "",
                    "min_qty": qty_min,
                    "max_qty": qty_max,
                    "unit_cost": parse_money(raw_price),
                    "setup_cost": transfer_setup,
                    "repeat_setup_cost": "",
                    "pricing_basis": "per_unit",
                    "price_status": price_status(raw_price),
                    "surcharge_note": "",
                    "source_row_number": row_idx,
                    "raw_price": raw_price,
                }
            )

    # Embroidery stitch-count matrix.
    embroidery_setup = "75.00"
    embroidery_qty_ranges = [clean(v) for v in next(ws.iter_rows(min_row=290, max_row=290, values_only=True))[1:10]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=291, max_row=306, values_only=True), start=291):
        stitch_min, stitch_max = parse_stitch_range(row[0])
        for offset, qty_range in enumerate(embroidery_qty_ranges, start=1):
            qty_min, qty_max = parse_qty_range(qty_range)
            raw_price = clean(row[offset])
            rows.append(
                {
                    "supplier": "Gear For Life",
                    "rate_card_key": "embroidery_apparel_selected_bags",
                    "decoration_method": "Embroidery",
                    "applies_to": "Apparel and selected bags",
                    "artwork_size_label": "",
                    "stitch_count_min": stitch_min,
                    "stitch_count_max": stitch_max,
                    "min_qty": qty_min,
                    "max_qty": qty_max,
                    "unit_cost": parse_money(raw_price),
                    "setup_cost": embroidery_setup,
                    "repeat_setup_cost": "",
                    "pricing_basis": "per_stitch_band",
                    "price_status": price_status(raw_price),
                    "surcharge_note": "Some items +$1.90/unit",
                    "source_row_number": row_idx,
                    "raw_price": raw_price,
                }
            )

    return cards, rows


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    ws = load_workbook_rows()
    options, price_rows, unavailable_rows = extract_product_specific_rows(ws)
    general_cards, general_rows = extract_general_rate_cards(ws)

    option_fields = [
        "supplier",
        "supplier_sku",
        "product_name",
        "decoration_option_key",
        "decoration_method",
        "decoration_area",
        "decoration_location",
        "artwork_size_label",
        "max_width_mm",
        "max_height_mm",
        "pricing_model",
        "price_status",
        "setup_cost",
        "repeat_setup_cost",
        "setup_cost_label",
        "pricing_basis",
        "additional_colour_policy",
        "source_row_number",
        "raw_json",
    ]
    price_fields = [
        "supplier",
        "supplier_sku",
        "product_name",
        "decoration_option_key",
        "decoration_method",
        "decoration_area",
        "decoration_location",
        "artwork_size_label",
        "currency",
        "min_qty",
        "max_qty",
        "unit_cost",
        "setup_cost",
        "repeat_setup_cost",
        "pricing_basis",
        "price_status",
        "raw_price",
        "source_row_number",
    ]
    general_fields = [
        "supplier",
        "rate_card_key",
        "decoration_method",
        "applies_to",
        "artwork_size_label",
        "stitch_count_min",
        "stitch_count_max",
        "min_qty",
        "max_qty",
        "unit_cost",
        "setup_cost",
        "repeat_setup_cost",
        "pricing_basis",
        "price_status",
        "surcharge_note",
        "source_row_number",
        "raw_price",
    ]
    general_card_fields = [
        "supplier",
        "rate_card_key",
        "decoration_method",
        "applies_to",
        "applies_to_category",
        "applies_to_subcategory",
        "is_default_for_scope",
        "fallback_policy",
        "pricing_model",
        "setup_cost",
        "repeat_setup_cost",
        "surcharge_cost",
        "surcharge_label",
        "notes",
    ]

    write_csv(OPTIONS_CSV, options, option_fields)
    write_csv(PRICE_ROWS_CSV, price_rows, price_fields)
    write_csv(GENERAL_RATE_CARDS_CSV, general_cards, general_card_fields)
    write_csv(GENERAL_RATE_ROWS_CSV, general_rows, general_fields)

    sku_count = len({r["supplier_sku"] for r in options if r["supplier_sku"]})
    method_counts = Counter(r["decoration_method"] for r in options)
    option_status_counts = Counter(r["price_status"] for r in options)
    price_status_counts = Counter(r["price_status"] for r in price_rows)
    general_status_counts = Counter(r["price_status"] for r in general_rows)

    audit = [
        "# Gear For Life Branding Workbook Audit",
        "",
        "Source workbook:",
        "",
        f"`{SOURCE_XLSX}`",
        "",
        "## Summary",
        "",
        f"- Product-specific SKUs: {sku_count}",
        f"- Product-specific decoration options: {len(options)}",
        f"- Product-specific decoration price rows: {len(price_rows)}",
        f"- General rate cards: {len(general_cards)}",
        f"- General rate card rows: {len(general_rows)}",
        f"- Option status counts: `{dict(option_status_counts)}`",
        f"- Price row status counts: `{dict(price_status_counts)}`",
        f"- General rate row status counts: `{dict(general_status_counts)}`",
        "",
        "## Method Counts",
        "",
    ]
    for method, count in method_counts.most_common():
        audit.append(f"- {method}: {count}")

    audit += [
        "",
        "## Special Handling",
        "",
        "- Rows with `POA` are preserved with `price_status = poa`; prices are not guessed.",
        "- Rows such as `Knife decoration is not applicable` are preserved with `price_status = unavailable`.",
        "- The Transfer Printing and Embroidery sections are general rate cards, not product-specific SKU rows.",
        "- Transfer Printing uses `rate_card_key = transfer_printing_bags` and applies to all Bags.",
        "- Transfer Printing is the default fallback for bag products when no product-specific decoration method is supplied.",
        "- Embroidery uses `rate_card_key = embroidery_apparel_selected_bags` and `pricing_basis = per_stitch_band`.",
        "- The import layer stores supplier cost only; margin belongs in the quote/pricing layer.",
        "",
        "## Output Preview Files",
        "",
        f"- `{OPTIONS_CSV.name}`",
        f"- `{PRICE_ROWS_CSV.name}`",
        f"- `{GENERAL_RATE_CARDS_CSV.name}`",
        f"- `{GENERAL_RATE_ROWS_CSV.name}`",
        "",
    ]

    if unavailable_rows:
        audit += ["## Unavailable Decoration Rows", ""]
        for row_idx, sku, location, raw_price in unavailable_rows:
            audit.append(f"- Row {row_idx}: {sku} / {location} / {raw_price}")

    AUDIT_MD.write_text("\n".join(audit) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "options": len(options),
                "price_rows": len(price_rows),
                "general_rate_cards": len(general_cards),
                "general_rate_rows": len(general_rows),
                "skus": sku_count,
                "option_status_counts": dict(option_status_counts),
                "price_status_counts": dict(price_status_counts),
                "general_status_counts": dict(general_status_counts),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
